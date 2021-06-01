#%%
import numpy as np
import csv
import string
import random
import openpyxl
from itertools import combinations

times_dict = {'08:15':0, '08:45':1,'09:15':2,'09:45':3, '10:15':4, '10:45':5,'11:15':6,'11:45':7, '12:15':8, '12:45':9,'13:15':10,'13:45':11, '14:15':12, '14:45':13,'15:15':14,'15:45':15, '16:15':16, '16:45':17,'17:15':18,'17:45':19, '18:15':20, '18:45':21,'19:15':22,'19:45':23, '20:15':24}
day_dict = {'M': 0, 'T': 1, 'W': 2, 'H': 3, 'F': 4}

duplicates_list = []

def make_2d_array():
    # Each row represents a day, each time represents a timeslot
    # Ex. Accessing Tuesday's 3rd timeslot: arr[1][2] (0-based indexing)
    return [[0] * 25 for i in range(5)]

class Course(object):
        # What is Course?
        # A COURSES_LIST will contain each course (likely for a specific subject).
    def __init__(self, title, course_no, section, times, room, teacher='N/A'):
        self.title = title
        self.course_no = course_no
        self.section = str(section).zfill(5)
        self.times = times        
        self.teacher = teacher
        # 'times' dict will look like this: 
            # Ex. M 8:15-10:15
            #     W 8:15-9:45
            #     F 14:15-15:45
            # {0: [0,4], 2: [0,3], 4:[12,15]}
        self.room = room
        self.course_type = self.get_course_type()
    
    def get_course_type(self):
        if self.course_no in eng_index:
            return 'ENG'
        elif self.course_no in hum_index:
            return 'HUM'
        elif self.course_no in com_index:
            return 'COM'
        elif self.course_no in lin_index:
            return 'LIN'
        elif self.course_no in bio_index:
            return 'BIO'
        elif self.course_no in wav_index:
            return 'WAV'
    
    def check_duplicate(self, Course1, Course2):
        check = False
        for day in Course1.times:
            try:
                check = Course1.times[day] == Course2.times[day]
            except:
                pass
        if check == True:
            duplicates_list.append((Course1, Course2))
        return check

    
    def __str__(self):
        #return "Title: " + str(self.title) + "\n" + "Course #: " + str(self.course_no) + "\n" + "Section: " + str(self.section) + "\n" + "Times: " + str(self.times) + "\n" + "Room: " + str(self.room) + "\n" + "Teacher: " + str(self.teacher) + "\n" + "Course type: " + self.course_type + "\n"
        return str(self.course_no + " " + self.section + " " + self.teacher)
    
    def __repr__(self):
        return str(self.course_no + " " + self.section + " " + self.teacher)

def csv_to_2d_arr(csv_name):
    with open(csv_name, 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        data = []
        for row in csv_reader:
            data.append(row)
        return data

def convert_table(arr):
    new_table = []
    index_1 = 0
    index_2 = 0
    for row in range(len(arr)-2):
        if arr[row][0] != '':
            index_1 = row
            if arr[row+1][0] != '':
                index_2 = row + 1
            elif arr[row+2][0] != '':
                index_2 = row + 2
            else:
                index_2 = row + 3
            new_table.append(arr[index_1:index_2])
    return new_table # Each row is a course

def string_to_time_list(string):
    time_list = []
    for time in string.split('-'):
        time_list.append(times_dict.get(time))
    return time_list 

def create_class_from_list(my_list): # Nested list
    section = my_list[0][0]
    course_no = my_list[0][1]
    room = my_list[0][5]
    title = ''
    teacher = ''
    times = {}
    for row in range(len(my_list)):
        lower_count = len([letter for letter in my_list[row][2] if letter in string.ascii_lowercase])
        if lower_count >= 2:
            teacher = my_list[row][2]
        else:
            title += " " + my_list[row][2]

        days = my_list[row][3]
        for day in days:
            new_entry = {int(day_dict.get(day)): string_to_time_list(my_list[row][4])}
            times.update(new_entry)
    return Course(title, course_no, section, times, room, teacher)


def print_2d_arr(arr):
    for row in arr:
        print(row)


####################

class Schedule(object):
    def __init__(self):
        self.timetable = make_2d_array()
        self.courses_taken = []
        self.subjects_taken = []
    
    def add_course(self, Course):
        for day in Course.times:
            index_0 = Course.times[day][0]
            index_1 = Course.times[day][1] 
            # self.timetable[day][index_0:index_1] = [Course] * (index_1 - index_0) 
            for slot in range(index_0, index_1):
                self.timetable[day][slot] = Course
        self.courses_taken.append(Course)
        self.subjects_taken.append(Course.course_type)


    def check_compatibility(self, Course):
        if not Course.course_type in self.subjects_taken: 
            for day in Course.times:
                index_0 = Course.times[day][0]
                index_1 = Course.times[day][1]
                if self.timetable[day][index_0] != 0 or self.timetable[day][index_1] != 0:
                    return False
                else:
                    return True
        else:
            return False
    
    def pick_random_schedule(self):
        failed_msg = "No working schedule was found, please try again."
        self.add_course(random.choice(eng_courses))
        for Course in hum_courses:
            if self.check_compatibility(Course):
                self.add_course(Course)
        if 'HUM' not in self.subjects_taken:
            return failed_msg
        for Course in com_courses:
            if self.check_compatibility(Course):
                self.add_course(Course)
        if 'COM' not in self.subjects_taken:
            return failed_msg
        for Course in lin_courses:
            if self.check_compatibility(Course):
                self.add_course(Course)
        if 'LIN' not in self.subjects_taken:
            return failed_msg
        for Course in bio_courses:
            if self.check_compatibility(Course):
                self.add_course(Course)
        if 'BIO' not in self.subjects_taken:
            return failed_msg
        for Course in wav_courses:
            if self.check_compatibility(Course):
                self.add_course(Course)
        if 'WAV' not in self.subjects_taken:
            return failed_msg
        #return self.timetable
    
    def delete_course(self, Course):
        for i, day in enumerate(self.timetable):
            for j, timeslot in enumerate(day):
                if repr(Course) == timeslot:
                    self.timetable[i][j] = 0

    
    def clear_timetable(self):
        self.timetable = make_2d_array()

    def find_all_timetables(self):
        timetables_list = []
        timetable_count = 0
        for bio_course in bio_courses:
            self.add_course(bio_course)
            for hum_course in hum_courses:
                if self.check_compatibility(hum_course):
                    self.add_course(hum_course)
                for lin_course in lin_courses:
                    if self.check_compatibility(lin_course):
                        self.add_course(lin_course)
                    for eng_course in eng_courses:
                        if self.check_compatibility(eng_course):
                            self.add_course(eng_course)
                        for wav_course in wav_courses:
                            if self.check_compatibility(wav_course):
                                self.add_course(wav_course)
                            for com_course in com_courses:
                                if self.check_compatibility(com_course):
                                    self.add_course(com_course)
                                if len(self.subjects_taken) == 6:
                                    timetables_list.append(self.timetable)
                                    timetable_count +=1 
                                    if timetable_count % 1000 == 0:
                                        print("Schedule found POGGIES" + str(timetable_count))
                                        print(self.timetable)
                                self.delete_course(com_course)
                            self.delete_course(wav_course)
                        self.delete_course(eng_course)
                    self.delete_course(lin_course)
                self.delete_course(hum_course)
            self.delete_course(bio_course)


##########################################

# 6 subjects -> 6 lists

eng_index = ['603-103-MQ']
hum_index = ['345-LPH-MS']
com_index = ['ARH-LAA', 'ART-LAA', 'ART-LBA', 'CIN-LAA', 'FLM-LBA', 'GER-LAL', 'GER-LBL', 'ITA-LAA', 'MAT-LAM', 'MUS-LAA', 'PHI-LAA', 'PRO-LAM', 'SSS-LAQ', 'SSS-LBQ', 'STS-LBT', 'THE-LAA'] 
lin_index = ['201-NYC-05']
bio_index = ['101-LCU-05']
wav_index = ['203-NYC-05']

eng_courses = []
hum_courses = []
com_courses = []
lin_courses = []
bio_courses = []
wav_courses = []

FULL_COURSES_LIST = convert_table(csv_to_2d_arr('full_courses.csv'))

#with open('courses_list.txt', 'w') as wf:
#    for row in FULL_COURSES_LIST:
#        wf.write(str(row) + "\n")

CONV_COURSES_LIST = [create_class_from_list(row) for row in FULL_COURSES_LIST]
#with open('courses_list2.txt', 'w') as wf:
#    for Course in CONV_COURSES_LIST:
#        wf.write(str(Course))

#for Course in CONV_COURSES_LIST[282:]:
#    if not Course.course_no in com_index:
#        com_index.append(Course.course_no)

for Course in CONV_COURSES_LIST:
    if Course.course_no in eng_index:
        eng_courses.append(Course)
    elif Course.course_no in hum_index:
        hum_courses.append(Course)
    elif Course.course_no in com_index:
        com_courses.append(Course)
    elif Course.course_no in lin_index:
        lin_courses.append(Course)
    elif Course.course_no in bio_index:
        bio_courses.append(Course)
    elif Course.course_no in wav_index:
        wav_courses.append(Course)

mySchedule = Schedule()

print("working")


#############################

wb = openpyxl.Workbook()
filepath = "courses.xlsx"
wb.save(filepath)

sheet = wb.active

def template(sheet):
    for day in day_dict:
        sheet.cell(row=1, column=day_dict[day]+2).value = day
    for time in times_dict:
        sheet.cell(row=times_dict[time]+2, column=1).value = time 

template(sheet)

subject_color_dict = {'ENG': '9803FC', 'HUM': 'FCAD03', 'COM': '03F8FC', 'LIN': 'FC0339', 'BIO': '03FC5E', 'WAV': 'FCF803'} 

def timetable_to_xlsx(timetable, col_adjust=0):
    for i, day in enumerate(timetable):
        for j, timeslot in enumerate(day):
            sheet.cell(row=j+2,column=i+col_adjust+2).value = str(timeslot)
            if timeslot != 0:
                sheet.cell(row=j+2,column=i+2).fill = openpyxl.styles.PatternFill(start_color=subject_color_dict[timeslot.course_type], fill_type="solid")

def timetables_to_xlsx(timetable_list):
    col_adjust = 0
    for timetable in timetable_list:
        timetable_to_xlsx(timetable, col_adjust)
        col_adjust += 7

def find_duplicates(my_list):
    duplicate_count = 0
    for comb in list(combinations(my_list, 2)):
        if comb[0].check_duplicate(comb[0], comb[1]):
            try:
                my_list.remove(comb[1])
            except:
                pass
            duplicate_count += 1
    print("Found " + str(duplicate_count) + " duplicates")


master_list = (eng_courses, hum_courses, com_courses, lin_courses, bio_courses, wav_courses)

def debug_courses():
    for course_list in master_list:
        print(len(course_list))

debug_courses()
for course_list in master_list:
    find_duplicates(course_list)
debug_courses()

def remove_course_by_keyword(keyword):
    for course_list in master_list:
        for course in course_list:
            if keyword in str(course):
                course_list.remove(course)

def keep_course_by_keyword(keyword):
    for course_list in master_list:
        for course in course_list:
            if keyword not in str(course):
                course_list.remove(course)

#def max_course_time(time):
#    for course_list in master_list:
#        for Course in course_list:
#            #for day in course.times:
#            if times_dict[time] in sum(Course.times.values()):
#                course_list.remove(Course)

user_input = ''
while user_input != 'done':
    input2 = input('Do you want to keep a teacher or remove a teacher? (type k/r): ')
    if input2 == 'r':
        remove_course_by_keyword(input('Please specify a teacher you do not want: '))
    elif input2 == 'k':
        keep_course_by_keyword(input("Please specify a favorite teacher you want to keep: "))
    debug_courses()
    user_input = input('Is that all? (type "done" if yes)')


timetables_to_xlsx(mySchedule.find_all_timetables())

wb.save(filepath)


# %%

